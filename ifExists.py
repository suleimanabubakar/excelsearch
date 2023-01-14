import contextlib
from openpyxl import Workbook
import openpyxl
from datetime import datetime




def search_next(id_no,target_workbook,target_column):
    wb = openpyxl.load_workbook(f'excels/{target_workbook}.xlsx')
    ws = wb.sheetnames

    for sheet in ws:
        wsheet = wb[sheet]
        for cell in wsheet[f'{target_column}']:
            with contextlib.suppress(Exception):
                # print(f'{cell.value} - {id_no}')
                if id_no == cell.value:
                    return "Found"
    return "Not Found"
        
        


def main_filter(present_workbook,target_workbook,present_column,target_column):
          
    wb = openpyxl.load_workbook(f'excels/{present_workbook}.xlsx')
    ws = wb.sheetnames


    all_availables = []
    do_not_exist = []

    for sheet in ws:
        wsheet = wb[sheet]
        for cell in wsheet[f'{present_column}']:
            id_no = cell.value

            is_found = search_next(id_no,target_workbook,target_column)

            print(f'*{id_no} IS {is_found.upper()}*')

            row_id = cell.row
            if is_found == "Found":
                all_availables.append(wsheet[row_id])
            else:
                do_not_exist.append(wsheet[row_id])
                
                




    
    n_wb = Workbook()
    new_sheet = n_wb.active




    for max_row,row in enumerate(all_availables,start=1):
        for max_col, cell in enumerate(row, start=1):
            new_sheet.cell(row=max_row,column=max_col).value=cell.value
    

    not_exist = n_wb.create_sheet('NOT AVAILABLE')
    for max_row,row in enumerate(do_not_exist,start=1):
        for max_col, cell in enumerate(row, start=1):
            not_exist.cell(row=max_row,column=max_col).value=cell.value




    resultFile = f'excels/SORTED_GENERATED_AT_{datetime.now()}.xlsx'
    n_wb.save(filename=resultFile)

    print('*RESULT FILE SUCCESSFULLY GENERATED*')






            
        
